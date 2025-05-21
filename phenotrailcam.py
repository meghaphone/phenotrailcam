# -*- coding: utf-8 -*-
"""
Created on Mon Apr 14 09:17:05 2025

@author: Anthony Boyle
"""

import glob
from PIL import Image, ExifTags
import shutil
import os
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

#Function to find the folder within directory
def find_folders(root_dir, folder_name):
    """
    Finds all folders with a specific name within a root directory.

    Args:
        root_dir (str): The path to the root directory to search in.
        folder_name (str): The name of the folder to find.

    Returns:
        list: A list of paths to the folders found.
    """
    matching_folders = []
    for root, dirs, files in os.walk(root_dir):
        if folder_name in dirs:
            for dir in dirs:
                if dir == folder_name:
                    matching_folders.append(os.path.join(root, dir))
    return matching_folders

# Function to get the DateTimeOriginal EXIF tag from an image
def get_image_date(image_path):
    try:
        exif_data = Image.open(image_path)._getexif()
        if exif_data:
            for tag, value in exif_data.items():
                if ExifTags.TAGS.get(tag) == 'DateTimeOriginal':
                    return value  # Expected format: "YYYY:MM:DD HH:MM:SS"
    except Exception as e:
        print(f"Could not process {image_path}: {e}")
    return None

# Function to parse the date string into a datetime object
def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y:%m:%d %H:%M:%S")
    except Exception as e:
        print(f"Error parsing date '{date_str}': {e}")
        return None

# Function to rename and sort all image files
def sort_and_rename(camera_name, image_dir):

    folder_paths = find_folders(image_dir,camera_name)
    print(folder_paths)
    #Prints statements if no folders are found
    if not folder_paths:
        print(f"No folders named '{camera_name}' found in {image_dir}")

    destination_path = os.path.join(image_dir, '..', 'Sorted_photos', camera_name)
    # Ensure destination folder exists
    if not os.path.exists(destination_path):
        print("Making destination folder: " + destination_path)
        os.mkdir(destination_path)

    for folder_path in folder_paths:
        num_images = 0
        print("Copying images from " + folder_path + " to " + destination_path)
        # Get list of image files in the specific folder
        file_list = glob.glob(os.path.join(folder_path, "*.[jJ][pP][gG]"))


        # Create a list of tuples: (datetime, file_path)
        file_date_list = []
        for file_path in file_list:
            date_time = get_image_date(file_path)
            if date_time:
                dt = parse_date(date_time)
                if dt:
                    file_date_list.append((dt, file_path))
                else:
                    print(f"No date information found for {file_path}")
    
        # Sort the list by the datetime object (chronologically)
        file_date_list.sort(key=lambda x: x[0])
    
        # Process and move each file in sorted order with new naming convention:
        # Format: IMG_####_YYYYMMDD_110000.jpg
        for idx, (dt, file_path) in enumerate(file_date_list, start=1):
            order_str = f"{idx:04d}"         # Four-digit sequence number
            date_str = dt.strftime("%Y%m%d")  # Date in YYYYMMDD format
            new_filename = f"IMG_{camera_name}_{date_str}_110000.jpg"
        
            # Define the destination path with the new filename
            destination_file_path = os.path.join(destination_path, new_filename)
        
            # Copy and rename the file
            shutil.copy(file_path, destination_file_path)
            num_images = num_images+1
            #print(f"Copied and renamed: {file_path} -> {destination_file_path}")
        print("Copied " + str(num_images) + " images")
        
    return None 

#Function that is used in calulating percentiles of color coordinates
def calculate_percentiles(values, percentiles):
    """Calculate specified percentiles for a list of values."""
    return [np.percentile(values, p) for p in percentiles]

#Function that find date from file name if sorted properly
def extract_image_metadata(image_path):
    """Extracts date, year, and DOY from the filename."""
    filename = os.path.basename(image_path)
    try:
        year = int(filename[11:15])
        month = int(filename[15:17])
        day = int(filename[17:19])
        file_date = datetime(year, month, day)
        doy = file_date.timetuple().tm_yday
        return file_date.strftime("%Y-%m-%d"), year, doy, file_date
    except (ValueError, IndexError) as e:
        raise ValueError(f"Invalid filename format: {filename}. Error: {e}")

#Extracts most image color values
def extract_avg_rgb(image_path, roi_box):
    """Calculate average RGB values within the ROI."""
    with Image.open(image_path) as img:
        img = img.convert("RGB")

        if roi_box:
            img = img.crop(roi_box)  # Apply cropping using the ROI box

        # Convert image to NumPy array
        pixels = np.array(img)
        
        # Ensure pixels is a NumPy array
        if not isinstance(pixels, np.ndarray):
            raise TypeError(f"Expected NumPy array, but got {type(pixels)}")

        # Split the image into RGB channels
        reds, greens, blues = pixels[:,:,0], pixels[:,:,1], pixels[:,:,2]
        
        # Ensure they're NumPy arrays (they should be after slicing)
        reds = reds.astype(np.float32)  # Convert to float for calculation
        greens = greens.astype(np.float32)
        blues = blues.astype(np.float32)

        total = reds + greens + blues
        gcc = np.divide(greens, total, out=np.zeros_like(greens, dtype=float), where=total != 0)
        rcc = np.divide(reds, total, out=np.zeros_like(reds, dtype=float), where=total != 0)

        # Calculate average values across the ROI
        avg_r = np.mean(reds[:])
        avg_g = np.mean(greens[:])
        avg_b = np.mean(blues[:])

        # Calculate standard deviations (over the pixel arrays)
        std_r = np.std(reds[:])
        std_g = np.std(greens[:])
        std_b = np.std(blues[:])

        # Calculate normalized average values (gcc, rcc)
        gcc_mean = np.mean(gcc)
        rcc_mean = np.mean(rcc[:])
        gcc_std = np.std(gcc[:])
        rcc_std = np.std(rcc[:])
        gcc_percentiles = calculate_percentiles(gcc[:], [50, 75, 90])
        rcc_percentiles = calculate_percentiles(rcc[:], [50, 75, 90])

        return {
            "avg_r": avg_r, "avg_g": avg_g, "avg_b": avg_b,
            "std_r": std_r, "std_g": std_g, "std_b": std_b,
            "gcc_mean": gcc_mean, "gcc_std": gcc_std,
            "gcc_50": gcc_percentiles[0], "gcc_75": gcc_percentiles[1], "gcc_90": gcc_percentiles[2],
            "rcc_mean": rcc_mean, "rcc_std": rcc_std,
            "rcc_50": rcc_percentiles[0], "rcc_75": rcc_percentiles[1], "rcc_90": rcc_percentiles[2],
        }

#Sets template for excel file, making all needed headers
def create_excel_file(filename, camera_name):
    """Creates the initial Excel file with metadata."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Metadata constants
    site_name = camera_name
    veg_type = "DB"
    roi_id_number = "1000"
    lat = "47"
    lon = "-90"
    elev = "160"
    utc_offset = "-6"
    img_count_threshold = "1"
    aggregation_period = "1"
    solar_elev_min = "10"
    tod_min = "00:00"
    tod_max = "23:59"
    roi_brightness_min = "100"
    roi_brightness_max = "665"
    creation_date = "2025-01-23"
    creation_time = "11:00"
    final_processing_date = "2025-01-25"
    final_processing_time = "13:00"

    #Setting flag values to 0
    snow_flag_val, outlierflag_gcc_mean, outlierflag_gcc_50, outlierflag_gcc_75, outlierflag_gcc_90 = 0, 0, 0, 0, 0
    
    # Write metadata rows (customize as needed)
    sheet.append(["#"])
    sheet.append(["# 1-day summary product time series for " + site_name])
    sheet.append(["#"])
    sheet.append(["# Site: " + site_name])
    sheet.append(["# Veg Type: " + veg_type])
    sheet.append(["# ROI ID Number: " + roi_id_number])
    sheet.append(["# Lat: " + lat])
    sheet.append(["# Lon: " + lon])
    sheet.append(["# Elev: " + elev])
    sheet.append(["# UTC Offset: " + utc_offset])
    sheet.append(["# Image Count Threshold: " + img_count_threshold])
    sheet.append(["# Aggregation Period: " + aggregation_period])
    sheet.append(["# Solar Elevation Min: " + solar_elev_min])
    sheet.append(["# Time of Day Min: " + tod_min])
    sheet.append(["# Time of Day Max: " + tod_max])
    sheet.append(["# ROI Brightness Min: " + roi_brightness_min])
    sheet.append(["# ROI Brightness Max: " + roi_brightness_max])
    sheet.append(["# Creation Date: " + creation_date])
    sheet.append(["# Creation Time: " + creation_time])
    sheet.append(["# Update Date: " + creation_date])
    sheet.append(["# Update Time: " + creation_time])
    sheet.append(["# Final Processing Date: " + final_processing_date])
    sheet.append(["# Final Processing Time: " + final_processing_time])
    sheet.append(["#"])
    # Write header row (this is row 25 in the final output)
    sheet.append(["date", "year", "doy", "image_count", "midday_filename", "midday_r", "midday_g", "midday_b",
                  "midday_gcc", "midday_rcc", "r_mean", "r_std", "g_mean", "g_std", "b_mean", "b_std",
                  "gcc_mean", "gcc_std", "gcc_50", "gcc_75", "gcc_90",
                  "rcc_mean", "rcc_std", "rcc_50", "rcc_75", "rcc_90", "max_solar_elev", "snow_flag", 
                  "outlierflag_gcc_mean", "outlierflag_gcc_50", "outlierflag_gcc_75", "outlierflag_gcc_90"])
    
    workbook.save(filename)
    print(f"Created Excel file: {filename}")
    
#Acquires mathcing roi and makes roi box needed for certain date
def get_roi(Camera_ID, date_str, roi_df):
    """Check for the matching ROI based on Camera ID and date."""
    image_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    #print(f"Checking ROIs for {Camera_ID} on {image_date}")

    # Filter the DataFrame for the relevant Camera ID
    filtered_df = roi_df[roi_df["Camera ID"] == Camera_ID]

    if filtered_df.empty:
        print(f"No ROI data found for Camera ID: {Camera_ID}")
        return None

    # Now loop through the filtered DataFrame for matching date ranges
    for _, row in filtered_df.iterrows():
        start_date = row["Start Date"].date()
        end_date = row["End Date"].date()
        
        # Convert the ROI coordinates to integers
        try:
            roi = (int(row["ROI top lft X"]), int(row["ROI top lft Y"]), 
                   int(row["ROI btm rt X"]), int(row["ROI btm rt Y"]))
        except ValueError as e:
            print(f"Error converting ROI coordinates: {e}")
            continue  # Skip this ROI if the conversion fails

        if start_date <= image_date <= end_date:
            #print(f"Match Found! ROI: {roi} for Camera ID: {Camera_ID} on {image_date}")
            return roi  # Return the ROI as a tuple

    print(f"No ROI match found for Camera ID: {Camera_ID} on {date_str}. Using default ROI.")
    return None

#Appends iamge data to excel file
def append_images_to_excel(Camera_ID, folder_path, excel_file, roi_ranges):
    """Extract RGB data from images and append to the Excel file."""
    if not os.path.exists(excel_file):
        create_excel_file(excel_file, Camera_ID)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    for filename in sorted(os.listdir(folder_path)):  
        if filename.lower().endswith((".png", ".jpg", ".jpeg", ".bmp")):
            image_path = os.path.join(folder_path, filename)
            try:
                # Extract date, year, doy from the image filename
                date_str, year, doy, file_date = extract_image_metadata(image_path)

                # Get the ROI box based on the Camera ID and date
                current_roi = get_roi(Camera_ID, date_str, roi_ranges)

                if current_roi is None:
                    print(f"Skipping image {filename} due to missing ROI.")
                    continue

                # Extract metrics using the selected ROI
                metrics = extract_avg_rgb(image_path, roi_box = current_roi)

                # Append extracted data to the sheet
                if abs(metrics["gcc_mean"] - (1/3)) >= 0.000001:#math.isclose(metrics["gcc_mean"], 0.33333, rel_tol=10**(-5): 
                                 sheet.append([date_str, year, doy, 1, filename,
                                    metrics["avg_r"], metrics["avg_g"], metrics["avg_b"],
                                    metrics["gcc_mean"], metrics["rcc_mean"], metrics["avg_r"], metrics["std_r"], 
                                    metrics["avg_g"], metrics["std_g"], metrics["avg_b"], metrics["std_b"],
                                    metrics["gcc_mean"], metrics["gcc_std"], metrics["gcc_50"], metrics["gcc_75"], 
                                    metrics["gcc_90"], metrics["rcc_mean"], metrics["rcc_std"], metrics["rcc_50"], 
                                    metrics["rcc_75"], metrics["rcc_90"], 0, 0, 0, 0, 0, 0]) 
            except Exception as e: 
                print(f"Error processing {filename}: {e}")

    workbook.save(excel_file)
    print(f"Data appended to {excel_file}")
    
#Finds dates that roi start and ends
def get_roi_ranges(csv_file):
    """Read the ROI data from CSV and create ROI date ranges."""
    roi_df = pd.read_csv(csv_file)

    # Check if the necessary columns exist
    required_columns = ["Starting Year", "Starting Month", "Starting Day", 
                        "Ending Year", "Ending Month", "Ending Day", 
                        "Camera ID", "ROI top lft X", "ROI top lft Y", 
                        "ROI btm rt X", "ROI btm rt Y"]

    # Ensure the required columns are in the CSV
    for col in required_columns:
        if col not in roi_df.columns:
            raise KeyError(f"Expected column '{col}' is missing from the CSV.")

    # Create 'Start Date' and 'End Date' by combining the year, month, and day columns
    roi_df["Start Date"] = pd.to_datetime(
        roi_df["Starting Year"].astype(str) + '-' +
        roi_df["Starting Month"].astype(str) + '-' +
        roi_df["Starting Day"].astype(str), errors="coerce"
    )

    roi_df["End Date"] = pd.to_datetime(
        roi_df["Ending Year"].astype(str) + '-' +
        roi_df["Ending Month"].astype(str) + '-' +
        roi_df["Ending Day"].astype(str), errors="coerce"
    )

    # Print out the first few rows to verify the dates
    #print("Start Date and End Date:\n", roi_df[["Start Date", "End Date"]].head())

    # Now sort the data by Start Date
    roi_df.sort_values(by="Start Date", inplace=True)

    # Return the dataframe with the created date ranges
    return roi_df

def set_flags(image_stats_file):

    # Read the entire Excel file without a header
    df_all = pd.read_excel(image_stats_file, header=None)

    # Assume that rows 1-25 (0-indexed rows 0-24) contain metadata (including the header row on row 25)
    metadata = df_all.iloc[:24, :]

    # Data rows start at row 26 (0-indexed row 25 and beyond)
    data = df_all.iloc[24:, :].copy()

    # Use the first row of the data portion (row 25 of the file) as the header for the data DataFrame
    data.columns = data.iloc[0]
    data = data.iloc[1:].reset_index(drop=True)

    # Ensure the "date" column is in datetime format
    data["date"] = pd.to_datetime(data["date"])

    # GCC_Outlier flag calculations
    top_thresholdmean = data["gcc_mean"].quantile(0.9)
    bottom_thresholdmean = data["gcc_mean"].quantile(0.1)

    top_threshold50 = data["gcc_50"].quantile(0.9)
    bottom_threshold50 = data["gcc_50"].quantile(0.1)

    top_threshold75 = data["gcc_75"].quantile(0.9)
    bottom_threshold75 = data["gcc_75"].quantile(0.1)

    top_threshold90 = data["gcc_90"].quantile(0.9)
    bottom_threshold90 = data["gcc_90"].quantile(0.1)

    # Flag Appending to file
    data["outlierflag_gcc_mean"] = np.where((data["gcc_mean"] >= top_thresholdmean) | (data["gcc_mean"] <= bottom_thresholdmean), 1, 0)
    data["outlierflag_gcc_50"] = np.where((data["gcc_50"] >= top_threshold50) | (data["gcc_50"] <= bottom_threshold50), 1, 0)
    data["outlierflag_gcc_75"] = np.where((data["gcc_75"] >= top_threshold75) | (data["gcc_75"] <= bottom_threshold75), 1, 0)
    data["outlierflag_gcc_90"] = np.where((data["gcc_90"] >= top_threshold90) | (data["gcc_90"] <= bottom_threshold90), 1, 0)

    # Snow flag
    data["snow_flag"] = np.where((data[["r_mean", "g_mean", "b_mean"]].max(axis=1) - 
                                  data[["r_mean", "g_mean", "b_mean"]].min(axis=1)) <= 1.5, 1, 0)

    # Extract month from date column
    data["month"] = data["date"].dt.month

    # Define the condition for snow and outlier flags
    flag_condition = (
        (data["snow_flag"] == 1) | 
        (data["outlierflag_gcc_mean"] == 1) | 
        (data["outlierflag_gcc_50"] == 1) | 
        (data["outlierflag_gcc_75"] == 1) | 
        (data["outlierflag_gcc_90"] == 1)
    )

    # Define the condition for October through March
    winter_months = data["month"].isin([11, 12, 1, 2])

    # Apply the replacement conditionally
    data.loc[flag_condition & winter_months, "gcc_90"] = np.nan
    data.loc[flag_condition & ~winter_months, "gcc_90"] = np.nan

    # Drop the temporary "month" column if not needed
    data.drop(columns=["month"], inplace=True)

    # Write the final result back to Excel while preserving the original metadata rows
    output_file = image_stats_file[:-5] +'_flag.xlsx'
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        metadata.to_excel(writer, index=False, header=False, startrow=0)
        data.to_excel(writer, index=False, startrow=24)

    print(f"Process completed. Check '{output_file}' for results.")
    return 

def clean_image_stats(input_file, output_file):
    #input_file = "CNAB02flag_DB_1000_1day.csv"      # Replace with your input CSV file path
    #output_file = "CNAB02flagclean_DB_1000_1day.csv"    # The new CSV file with cleaned data

    header_row_index = 24  # 0-indexed: row 25 is index 24

    with open(input_file, 'r') as infile, open(output_file, 'w') as outfile:
        for i, line in enumerate(infile):
            if i < header_row_index:
                # Remove all commas from rows before the header row
                cleaned_line = line.replace(",", "")
            else:
                cleaned_line = line
            outfile.write(cleaned_line)

    print("Finished cleaning CSV file.")
    return
